import tkinter as tk

# # Create window Tkinter
# window = tk.Tk()

# window.title(" DNAC Process Template Execution ")

# # window.geometry("700x500")

# frame = tk.Frame(window)
# frame.pack()

# # Saving User Info
# user_info_frame = tk.LabelFrame(frame, text="User Information")
# user_info_frame.grid(row=0, column=0)

# first_name_label = tk.Label(user_info_frame, text="First Name")
# first_name_label.grid(row=0, column=0)
# last_name_label = tk.Label(user_info_frame, text="Last Name")
# last_name_label.grid(row=0, column=1)

# first_name_entry = tk.Entry(user_info_frame)
# last_name_entry = tk.Entry(user_info_frame)
# first_name_entry.grid(row=1, column=0)
# last_name_entry.grid(row=1, column=1)

# window.mainloop()


import tkinter as tk
from tkinter import filedialog
import openpyxl

def upload_document():
    file_path = filedialog.askopenfilename(filetypes=[("XLSM Files", "*.xlsm")])
    if file_path:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)
            sheet = workbook["ENG-Script"]
            devices = []
            pre_implementation_commands = []
            post_implementation_commands = []
            start_capturing_devices = False
            start_capturing_pre_cmds = False
            start_capturing_post_cmds = False

            for row in sheet.iter_rows(values_only=True):
                if "PRE IMPLEMENTATION STEPS" in row:
                    start_capturing_devices = True
                    continue
                if start_capturing_devices and row[1] is not None:
                    devices.append(row[1])
                if "Step 1.0 D2O – Alarms will not be suppressed.  " in row:
                    start_capturing_devices = False

            for row in sheet.iter_rows(values_only=True):
                if "Step 1.2 D2O - Backup, issue baseline commands." in row:
                    start_capturing_pre_cmds = True
                    continue
                if start_capturing_pre_cmds and row[1] is not None:
                    pre_implementation_commands.append(row[1])
                if "Step 1.3 D2O – Prior to start on the maintenance, D2O contacts Carrier to confirm that maintenance is still scheduled." in row:
                    start_capturing_pre_cmds = False

            for row in sheet.iter_rows(values_only=True):
                if "Step 2.3  D2O – Verify circuit is up and run baseline commands and compare to pre-maintenance baselines." in row:
                    start_capturing_post_cmds = True
                    continue
                if start_capturing_post_cmds and row[1] is not None:
                    post_implementation_commands.append(row[1])
                if "Step 2.4  D2O – Release Code Yellow on redundant link, test (including extended ping test)" in row:
                    start_capturing_post_cmds = False

            devices = devices[:-1]  
            pre_implementation_commands = pre_implementation_commands[:-1]  
            post_implementation_commands = post_implementation_commands[:-1]  

            print("\n Devices: ", devices)
            print("\n Pre-Implementation Commands: ", pre_implementation_commands)
            print("\n Post-Implementation Commands: ", post_implementation_commands)

        except Exception as e:
            print("Error reading the XLSM file:", e)

# Create the main window
window = tk.Tk()
window.title("NCD Reader")
window.geometry("400x400")

# Create GUI components
label = tk.Label(window, text="Select a NCD Document:")
button = tk.Button(window, text="Upload", command=upload_document)

# Add GUI components to the window 
label.pack(padx=10, pady=10)
button.pack()

# Start the application's event loop
window.mainloop()
